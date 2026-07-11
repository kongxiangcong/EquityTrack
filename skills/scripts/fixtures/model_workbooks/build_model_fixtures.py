#!/usr/bin/env python3
"""Build minimal workbook fixtures for model_validator.py."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook


ROOT = Path(__file__).resolve().parent
OUT = ROOT / "generated"


YEARS = ["FY-4A", "FY-3A", "FY-2A", "FY-1A", "FY0A", "FY+1E", "FY+2E", "FY+3E", "FY+4E", "FY+5E"]


def reset_workbook() -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    for name in [
        "Raw Data",
        "Operating Drivers",
        "Revenue Model",
        "Income Statement",
        "Balance Sheet",
        "Cash Flow",
        "DCF",
        "Comps",
        "Sensitivity",
    ]:
        wb.create_sheet(name)
    return wb


def fill_headers(ws, title: str) -> None:
    ws["A1"] = title
    ws["B3"] = "Line Item"
    for index, year in enumerate(YEARS, start=3):
        ws.cell(row=3, column=index).value = year


def fill_raw_data(ws) -> None:
    fill_headers(ws, "FixtureCo Raw Data")
    ws["M3"] = "source_id"
    rows = {
        6: "Revenue",
        7: "Net Income",
        8: "EPS",
        28: "Cash",
        29: "Debt",
        36: "Total Assets",
        45: "Total Liabilities & Equity",
        60: "Cash From Operations",
        61: "Capital Expenditure",
        62: "Free Cash Flow",
        63: "Net Change in Cash",
        64: "Beginning Cash",
        65: "Ending Cash",
    }
    for row, label in rows.items():
        ws.cell(row=row, column=2).value = label
        for col in range(3, 8):
            ws.cell(row=row, column=col).value = 100 + row + col
        ws.cell(row=row, column=13).value = "SRC-FILING-2024"


def fill_operating_drivers(ws) -> None:
    fill_headers(ws, "FixtureCo Operating Drivers")
    ws["M3"] = "Source"
    ws["N3"] = "Rationale"
    for row, label in {
        6: "Revenue Growth Drivers",
        8: "Gross Margin %",
        12: "Tax Rate",
        19: "CapEx / Revenue %",
        28: "Diluted Share Count",
    }.items():
        ws.cell(row=row, column=2).value = label
        for col in range(3, 13):
            ws.cell(row=row, column=col).value = 0.1 if row != 28 else 100
        ws.cell(row=row, column=13).value = "SRC-FILING-2024"
        ws.cell(row=row, column=14).value = "Historical average / source-gated fixture."


def fill_revenue_model(ws) -> None:
    fill_headers(ws, "FixtureCo Revenue Model")
    ws["B10"] = "Total Revenue"
    for col in range(3, 8):
        ws.cell(row=10, column=col).value = f"='Raw Data'!{ws.cell(row=1, column=col).coordinate}".replace("1", "6")
    for col in range(8, 13):
        prev = ws.cell(row=10, column=col - 1).coordinate
        ws.cell(row=10, column=col).value = f"={prev}*(1+'Operating Drivers'!{ws.cell(row=6, column=col).coordinate})"


def fill_income_statement(ws) -> None:
    fill_headers(ws, "FixtureCo Income Statement")
    line_rows = {6: "Revenue", 7: "Cost of Revenue", 8: "Gross Profit", 12: "Operating Income", 18: "Net Income", 20: "Diluted EPS"}
    for row, label in line_rows.items():
        ws.cell(row=row, column=2).value = label
    for col in range(3, 8):
        ws.cell(row=6, column=col).value = f"='Raw Data'!{ws.cell(row=6, column=col).coordinate}"
        ws.cell(row=18, column=col).value = f"='Raw Data'!{ws.cell(row=7, column=col).coordinate}"
        ws.cell(row=20, column=col).value = f"='Raw Data'!{ws.cell(row=8, column=col).coordinate}"
        ws.cell(row=7, column=col).value = f"={ws.cell(row=6, column=col).coordinate}*0.5"
        ws.cell(row=8, column=col).value = f"={ws.cell(row=6, column=col).coordinate}-{ws.cell(row=7, column=col).coordinate}"
        ws.cell(row=12, column=col).value = f"={ws.cell(row=8, column=col).coordinate}*0.6"
    for col in range(8, 13):
        ws.cell(row=6, column=col).value = f"='Revenue Model'!{ws.cell(row=10, column=col).coordinate}"
        ws.cell(row=7, column=col).value = f"={ws.cell(row=6, column=col).coordinate}*(1-'Operating Drivers'!{ws.cell(row=8, column=col).coordinate})"
        ws.cell(row=8, column=col).value = f"={ws.cell(row=6, column=col).coordinate}-{ws.cell(row=7, column=col).coordinate}"
        ws.cell(row=12, column=col).value = f"={ws.cell(row=8, column=col).coordinate}*0.6"
        ws.cell(row=18, column=col).value = f"={ws.cell(row=12, column=col).coordinate}*(1-'Operating Drivers'!{ws.cell(row=12, column=col).coordinate})"
        ws.cell(row=20, column=col).value = f"={ws.cell(row=18, column=col).coordinate}/'Operating Drivers'!{ws.cell(row=28, column=col).coordinate}"


def fill_balance_sheet(ws) -> None:
    fill_headers(ws, "FixtureCo Balance Sheet")
    for row, label in {6: "Cash", 10: "Total Assets", 20: "Total Liabilities & Equity", 22: "BALANCE CHECK"}.items():
        ws.cell(row=row, column=2).value = label
    for col in range(3, 8):
        ws.cell(row=6, column=col).value = f"='Raw Data'!{ws.cell(row=28, column=col).coordinate}"
        ws.cell(row=10, column=col).value = f"='Raw Data'!{ws.cell(row=36, column=col).coordinate}"
        ws.cell(row=20, column=col).value = f"='Raw Data'!{ws.cell(row=45, column=col).coordinate}"
    for col in range(8, 13):
        ws.cell(row=6, column=col).value = f"='Cash Flow'!{ws.cell(row=12, column=col).coordinate}"
        ws.cell(row=10, column=col).value = f"={ws.cell(row=6, column=col).coordinate}+500"
        ws.cell(row=20, column=col).value = f"={ws.cell(row=10, column=col).coordinate}"
    for col in range(3, 13):
        ws.cell(row=22, column=col).value = 0


def fill_cash_flow(ws) -> None:
    fill_headers(ws, "FixtureCo Cash Flow")
    for row, label in {
        6: "Cash From Operations",
        7: "Capital Expenditure",
        8: "Free Cash Flow",
        10: "Net Change in Cash",
        11: "Beginning Cash",
        12: "Ending Cash",
        14: "CASH TIE-OUT",
    }.items():
        ws.cell(row=row, column=2).value = label
    for col in range(3, 8):
        ws.cell(row=6, column=col).value = f"='Raw Data'!{ws.cell(row=60, column=col).coordinate}"
        ws.cell(row=7, column=col).value = f"='Raw Data'!{ws.cell(row=61, column=col).coordinate}"
        ws.cell(row=8, column=col).value = f"='Raw Data'!{ws.cell(row=62, column=col).coordinate}"
        ws.cell(row=10, column=col).value = f"='Raw Data'!{ws.cell(row=63, column=col).coordinate}"
        ws.cell(row=11, column=col).value = f"='Raw Data'!{ws.cell(row=64, column=col).coordinate}"
        ws.cell(row=12, column=col).value = f"='Raw Data'!{ws.cell(row=65, column=col).coordinate}"
    for col in range(8, 13):
        ws.cell(row=6, column=col).value = f"='Income Statement'!{ws.cell(row=18, column=col).coordinate}+10"
        ws.cell(row=7, column=col).value = f"='Income Statement'!{ws.cell(row=6, column=col).coordinate}*'Operating Drivers'!{ws.cell(row=19, column=col).coordinate}"
        ws.cell(row=8, column=col).value = f"={ws.cell(row=6, column=col).coordinate}-{ws.cell(row=7, column=col).coordinate}"
        ws.cell(row=10, column=col).value = f"={ws.cell(row=8, column=col).coordinate}"
        ws.cell(row=11, column=col).value = f"='Balance Sheet'!{ws.cell(row=6, column=col - 1).coordinate}"
        ws.cell(row=12, column=col).value = f"={ws.cell(row=11, column=col).coordinate}+{ws.cell(row=10, column=col).coordinate}"
    for col in range(3, 13):
        ws.cell(row=14, column=col).value = 0


def fill_dcf(ws, wacc: float = 0.10, terminal_growth: float = 0.03) -> None:
    ws["A1"] = "DCF"
    ws["B5"] = "WACC"
    ws["C5"] = wacc
    ws["B6"] = "Terminal Growth Rate"
    ws["C6"] = terminal_growth
    ws["B10"] = "Unlevered Free Cash Flow (FCFF)"
    ws["C10"] = "='Cash Flow'!H8"
    ws["B17"] = "Terminal Value"
    ws["C17"] = "=C10*(1+C6)/(C5-C6)"
    ws["B20"] = "Equity Bridge"
    ws["B21"] = "Enterprise Value"
    ws["B25"] = "Equity Value per Share"
    ws["C25"] = 12.3


def fill_comps(ws) -> None:
    ws["A1"] = "Comps"
    ws["A3"] = "Company"
    ws["B3"] = "P/E"
    for row, name in enumerate(["Peer A", "Peer B", "Peer C", "Target Company"], start=4):
        ws.cell(row=row, column=1).value = name
        ws.cell(row=row, column=2).value = 10 + row


def fill_sensitivity(ws) -> None:
    ws["A1"] = "Sensitivity"
    ws["B5"] = "DCF Base Case Value"
    ws["C5"] = 12.3
    ws["B6"] = "Sensitivity Center Check"
    ws["C6"] = 0
    ws["B10"] = "BASE"


def build_valid(wacc: float = 0.10, terminal_growth: float = 0.03) -> Workbook:
    wb = reset_workbook()
    fill_raw_data(wb["Raw Data"])
    fill_operating_drivers(wb["Operating Drivers"])
    fill_revenue_model(wb["Revenue Model"])
    fill_income_statement(wb["Income Statement"])
    fill_balance_sheet(wb["Balance Sheet"])
    fill_cash_flow(wb["Cash Flow"])
    fill_dcf(wb["DCF"], wacc=wacc, terminal_growth=terminal_growth)
    fill_comps(wb["Comps"])
    fill_sensitivity(wb["Sensitivity"])
    return wb


def main() -> None:
    OUT.mkdir(parents=True, exist_ok=True)

    build_valid().save(OUT / "valid_workbook.xlsx")

    missing = build_valid()
    missing.remove(missing["Cash Flow"])
    missing.save(OUT / "missing_required_sheet.xlsx")

    build_valid(wacc=0.02, terminal_growth=0.03).save(OUT / "wacc_le_terminal_growth.xlsx")

    build_valid().save(OUT / "dcf_disabled_output_exists.xlsx")
    build_valid().save(OUT / "financial_firm_fcff_wacc_dcf.xlsx")

    print(f"Wrote fixtures to {OUT}")


if __name__ == "__main__":
    main()
