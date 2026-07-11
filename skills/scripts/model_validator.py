#!/usr/bin/env python3
"""Validate equity research financial model workbooks.

This validator is intentionally workbook-level. It does not try to rebuild a
model, fetch market data, or calculate a valuation. Instead it checks that the
Excel artifact contains the required control surfaces, source traceability, and
method-specific guardrails before a full report is allowed.

Usage:
    python model_validator.py --workbook path/to/model.xlsx --pretty
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Any, Dict, Iterable, List, Mapping, Optional, Sequence, Tuple

try:
    from openpyxl import load_workbook  # type: ignore
    from openpyxl.utils import get_column_letter  # type: ignore
except ImportError:  # pragma: no cover - depends on local environment
    load_workbook = None
    get_column_letter = None


VALID_DCF_STATUSES = {"auto", "allowed", "caution", "disabled", "not_selected"}
VALID_COMPANY_TYPES = {
    "general",
    "financial",
    "bank",
    "insurance",
    "broker",
    "biopharma",
    "pre_revenue_biopharma",
    "innovative_drug",
}

BASE_REQUIRED_SHEETS = (
    "Raw Data",
    "Operating Drivers",
    "Revenue Model",
    "Income Statement",
    "Balance Sheet",
    "Cash Flow",
    "Comps",
)

HISTORICAL_TRACE_SHEETS = ("Revenue Model", "Income Statement", "Balance Sheet", "Cash Flow")
FORECAST_FORMULA_SHEETS = ("Revenue Model", "Income Statement", "Balance Sheet", "Cash Flow")
CHECK_TOLERANCE = 1.0

RAW_DATA_LINE_PATTERNS = (
    r"revenue",
    r"sales",
    r"net income",
    r"eps",
    r"cash",
    r"debt",
    r"total assets",
    r"total liabilities",
    r"equity",
    r"cfo|cash from operations|operating cash flow",
    r"capex|capital expenditure",
    r"fcf|free cash flow",
)

MODEL_LINE_PATTERNS = (
    r"revenue",
    r"cogs|cost of revenue",
    r"gross profit",
    r"operating income|ebit",
    r"net income",
    r"eps",
    r"cash",
    r"accounts receivable|ar",
    r"inventory",
    r"total assets",
    r"debt",
    r"total liabilities",
    r"total equity",
    r"cash from operations|cfo",
    r"capital expenditure|capex",
    r"free cash flow|fcf",
    r"ending cash",
)

DCF_COMPONENT_PATTERNS = {
    "fcff": (r"fcff", r"ufcf", r"unlevered free cash flow", r"free cash flow"),
    "wacc": (r"\bwacc\b", r"weighted average cost of capital"),
    "terminal_value": (r"terminal value", r"gordon"),
    "equity_bridge": (r"equity bridge", r"enterprise value", r"equity value"),
    "per_share": (r"per share", r"diluted shares", r"equity value per share"),
}

TARGET_PRICE_PATTERNS = (
    r"target price",
    r"price target",
    r"目标价",
    r"implied price",
    r"implied value per share",
    r"equity value per share",
    r"dcf value per share",
    r"upside",
    r"downside",
)


@dataclass
class Issue:
    severity: str
    code: str
    message: str
    path: str
    details: Dict[str, Any] = field(default_factory=dict)


class ModelValidator:
    def __init__(
        self,
        workbook_formula: Any,
        workbook_values: Any,
        workbook_path: Path,
        dcf_status: str = "auto",
        company_type: str = "general",
    ) -> None:
        self.wb = workbook_formula
        self.values_wb = workbook_values
        self.workbook_path = workbook_path
        self.dcf_status = self.resolve_dcf_status(dcf_status)
        self.company_type = normalize_token(company_type)
        self.issues: List[Issue] = []
        self.required_sheets = list(BASE_REQUIRED_SHEETS)
        if self.dcf_status in {"allowed", "caution"}:
            self.required_sheets.append("DCF")

    def validate(self) -> Dict[str, Any]:
        self.check_required_sheets()
        self.check_raw_data_source_coverage()
        self.check_historical_traceability()
        self.check_balance_sheet_balances()
        self.check_cash_flow_ties()
        self.check_forecast_cells_are_formula_driven()
        self.check_dcf_rules()
        self.check_company_type_rules()
        return self.result()

    def check_required_sheets(self) -> None:
        for sheet_name in self.required_sheets:
            if not self.has_sheet(sheet_name):
                self.add_issue(
                    "error",
                    "REQUIRED_SHEET_MISSING",
                    f"Required workbook sheet is missing: {sheet_name}.",
                    f"$.sheets.{sheet_name}",
                )

    def check_raw_data_source_coverage(self) -> None:
        ws = self.sheet("Raw Data")
        if ws is None:
            return

        source_columns = self.find_source_columns(ws)
        if not source_columns:
            self.add_issue(
                "error",
                "RAW_DATA_SOURCE_ID_COLUMN_MISSING",
                "Raw Data must include a source_id/source column or source map.",
                "$.sheets.Raw Data",
            )
            return

        missing_rows = []
        for row in range(1, ws.max_row + 1):
            label = row_label(ws, row)
            if not matches_any(label, RAW_DATA_LINE_PATTERNS):
                continue
            if not row_has_numeric_data(ws, row, excluded_columns=source_columns):
                continue
            if not row_has_source_id(ws, row, source_columns):
                missing_rows.append(row)

        if missing_rows:
            self.add_issue(
                "error",
                "RAW_DATA_SOURCE_ID_MISSING",
                "Raw Data rows with critical financial data must carry source_id coverage.",
                "$.sheets.Raw Data",
                {"rows": missing_rows[:20], "missing_row_count": len(missing_rows)},
            )

    def check_historical_traceability(self) -> None:
        for sheet_name in HISTORICAL_TRACE_SHEETS:
            ws = self.sheet(sheet_name)
            if ws is None:
                continue

            historical_columns = period_columns(ws, historical=True)
            if not historical_columns:
                historical_columns = [col for col in range(3, min(ws.max_column, 7) + 1)]

            missing_refs = []
            for row in range(1, ws.max_row + 1):
                label = row_label(ws, row)
                if not matches_any(label, MODEL_LINE_PATTERNS):
                    continue
                if is_check_or_tie_label(label):
                    continue
                for col in historical_columns:
                    cell = ws.cell(row=row, column=col)
                    if is_blank(cell.value):
                        continue
                    if is_formula(cell.value):
                        continue
                    if numeric_value(cell.value) is None:
                        continue
                    if not formula_references_sheet(cell.value, "Raw Data"):
                        missing_refs.append(cell_ref(ws, row, col))

            if missing_refs:
                self.add_issue(
                    "error",
                    "HISTORICAL_DATA_NOT_RAW_DATA_LINKED",
                    f"{sheet_name} historical financial cells must trace to Raw Data formulas.",
                    f"$.sheets.{sheet_name}",
                    {"cells": missing_refs[:25], "cell_count": len(missing_refs)},
                )

    def check_balance_sheet_balances(self) -> None:
        ws = self.sheet("Balance Sheet")
        values_ws = self.values_sheet("Balance Sheet")
        if ws is None or values_ws is None:
            return

        check_row = find_row(ws, (r"balance check", r"assets\s*-\s*(liabilities|l\+e)"))
        if check_row is not None:
            self.check_zero_row(values_ws, check_row, "BALANCE_SHEET_OUT_OF_BALANCE", "Balance Sheet balance check")
            return

        assets_row = find_row(ws, (r"total assets",))
        le_row = find_row(ws, (r"total liabilities.*equity", r"total l\s*\+\s*e", r"total liabilities\s*&\s*equity"))
        if assets_row is None or le_row is None:
            self.add_issue(
                "error",
                "BALANCE_CHECK_ROW_MISSING",
                "Balance Sheet must include a balance check row or comparable total assets / total liabilities plus equity rows.",
                "$.sheets.Balance Sheet",
            )
            return

        bad_columns = []
        for col in value_columns(ws):
            assets = numeric_value(values_ws.cell(row=assets_row, column=col).value)
            le = numeric_value(values_ws.cell(row=le_row, column=col).value)
            if assets is None or le is None:
                continue
            if abs(assets - le) > CHECK_TOLERANCE:
                bad_columns.append({"column": column_letter(col), "difference": assets - le})

        if bad_columns:
            self.add_issue(
                "error",
                "BALANCE_SHEET_OUT_OF_BALANCE",
                "Balance Sheet total assets must equal total liabilities plus equity.",
                "$.sheets.Balance Sheet",
                {"columns": bad_columns[:20], "bad_column_count": len(bad_columns)},
            )

    def check_cash_flow_ties(self) -> None:
        ws = self.sheet("Cash Flow")
        values_ws = self.values_sheet("Cash Flow")
        if ws is None or values_ws is None:
            return

        tie_row = find_row(ws, (r"cash tie", r"cash tie-out", r"ending cash.*bs cash"))
        if tie_row is not None:
            self.check_zero_row(values_ws, tie_row, "CASH_FLOW_TIE_OUT_FAILED", "Cash Flow cash tie-out")
            return

        ending_row = find_row(ws, (r"ending cash",))
        net_change_row = find_row(ws, (r"net change.*cash",))
        beginning_row = find_row(ws, (r"beginning cash",))
        if ending_row is None or net_change_row is None or beginning_row is None:
            self.add_issue(
                "error",
                "CASH_FLOW_TIE_OUT_ROW_MISSING",
                "Cash Flow must include cash tie-out or beginning/net-change/ending cash rows.",
                "$.sheets.Cash Flow",
            )
            return

        bad_columns = []
        for col in value_columns(ws):
            beginning = numeric_value(values_ws.cell(row=beginning_row, column=col).value)
            net_change = numeric_value(values_ws.cell(row=net_change_row, column=col).value)
            ending = numeric_value(values_ws.cell(row=ending_row, column=col).value)
            if beginning is None or net_change is None or ending is None:
                continue
            if abs((beginning + net_change) - ending) > CHECK_TOLERANCE:
                bad_columns.append({"column": column_letter(col), "difference": (beginning + net_change) - ending})

        if bad_columns:
            self.add_issue(
                "error",
                "CASH_FLOW_CASH_CHANGE_DOES_NOT_TIE",
                "Cash Flow ending cash must equal beginning cash plus net cash change.",
                "$.sheets.Cash Flow",
                {"columns": bad_columns[:20], "bad_column_count": len(bad_columns)},
            )

    def check_zero_row(self, values_ws: Any, row: int, code: str, label: str) -> None:
        bad_columns = []
        not_evaluable = []
        for col in value_columns(values_ws):
            value = numeric_value(values_ws.cell(row=row, column=col).value)
            if value is None:
                formula_ws = self.sheet(values_ws.title)
                formula_value = formula_ws.cell(row=row, column=col).value if formula_ws is not None else None
                if not is_blank(formula_value):
                    not_evaluable.append(cell_ref(values_ws, row, col))
                continue
            if abs(value) > CHECK_TOLERANCE:
                bad_columns.append({"cell": cell_ref(values_ws, row, col), "value": value})

        if bad_columns:
            self.add_issue(
                "error",
                code,
                f"{label} must be zero for every period.",
                f"$.sheets.{values_ws.title}",
                {"cells": bad_columns[:20], "bad_cell_count": len(bad_columns)},
            )
        if not_evaluable:
            self.add_issue(
                "warning",
                "CHECK_ROW_NOT_EVALUABLE",
                f"{label} contains formulas without cached values; open/save in Excel or provide visible check values for validation.",
                f"$.sheets.{values_ws.title}",
                {"cells": not_evaluable[:20], "cell_count": len(not_evaluable)},
            )

    def check_forecast_cells_are_formula_driven(self) -> None:
        for sheet_name in FORECAST_FORMULA_SHEETS:
            ws = self.sheet(sheet_name)
            if ws is None:
                continue
            forecast_columns = period_columns(ws, historical=False)
            if not forecast_columns:
                forecast_columns = [col for col in range(8, min(ws.max_column, 12) + 1)]

            hardcoded_cells = []
            for row in range(1, ws.max_row + 1):
                label = row_label(ws, row)
                if not matches_any(label, MODEL_LINE_PATTERNS):
                    continue
                if is_check_or_tie_label(label):
                    continue
                for col in forecast_columns:
                    cell = ws.cell(row=row, column=col)
                    if is_blank(cell.value):
                        continue
                    if is_formula(cell.value):
                        continue
                    if numeric_value(cell.value) is not None:
                        hardcoded_cells.append(cell_ref(ws, row, col))

            if hardcoded_cells:
                self.add_issue(
                    "error",
                    "FORECAST_CELL_HARDCODED",
                    f"{sheet_name} forecast cells must be formulas linked to drivers/model tabs, not hardcoded numbers.",
                    f"$.sheets.{sheet_name}",
                    {"cells": hardcoded_cells[:25], "cell_count": len(hardcoded_cells)},
                )

    def check_dcf_rules(self) -> None:
        dcf_ws = self.sheet("DCF")
        if self.dcf_status in {"allowed", "caution"}:
            if dcf_ws is None:
                self.add_issue(
                    "error",
                    "DCF_REQUIRED_SHEET_MISSING",
                    "DCF is allowed/caution, so the workbook must include a DCF sheet.",
                    "$.sheets.DCF",
                )
                return

            sheet_text = worksheet_text(dcf_ws)
            missing_components = [
                component
                for component, patterns in DCF_COMPONENT_PATTERNS.items()
                if not matches_any(sheet_text, patterns)
            ]
            if missing_components:
                self.add_issue(
                    "error",
                    "DCF_COMPONENT_MISSING",
                    "DCF sheet is missing required FCFF/WACC/terminal value/equity bridge/sensitivity components.",
                    "$.sheets.DCF",
                    {"missing_components": missing_components},
                )

            wacc = find_label_numeric_value(self.values_sheet("DCF"), (r"\bwacc\b", r"weighted average cost of capital"))
            terminal_growth = find_label_numeric_value(
                self.values_sheet("DCF"),
                (r"terminal growth", r"terminal growth rate", r"\bg\b"),
                reject_patterns=(r"terminal value",),
            )
            if wacc is None or terminal_growth is None:
                self.add_issue(
                    "error",
                    "DCF_WACC_OR_G_MISSING",
                    "DCF allowed/caution requires numeric WACC and terminal growth inputs.",
                    "$.sheets.DCF",
                    {"wacc": wacc, "terminal_growth": terminal_growth},
                )
            elif wacc <= terminal_growth:
                self.add_issue(
                    "error",
                    "DCF_WACC_NOT_GREATER_THAN_G",
                    "DCF terminal value requires WACC greater than terminal growth.",
                    "$.sheets.DCF",
                    {"wacc": wacc, "terminal_growth": terminal_growth},
                )

            self.check_sensitivity_center()
        elif dcf_ws is not None and sheet_contains_any(dcf_ws, TARGET_PRICE_PATTERNS):
            self.add_issue(
                "error",
                "DCF_DISABLED_OUTPUT_EXISTS",
                "DCF is disabled/not selected, but the workbook contains DCF target-price-style outputs.",
                "$.sheets.DCF",
                {"dcf_status": self.dcf_status},
            )

    def check_sensitivity_center(self) -> None:
        sensitivity_ws = self.sheet("Sensitivity")
        sensitivity_values_ws = self.values_sheet("Sensitivity")
        if sensitivity_ws is None or sensitivity_values_ws is None:
            self.add_issue(
                "error",
                "SENSITIVITY_SHEET_MISSING",
                "DCF allowed/caution requires a Sensitivity sheet with a base-case center check.",
                "$.sheets.Sensitivity",
            )
            return

        check_row = find_row(sensitivity_ws, (r"sensitivity.*center.*check", r"center.*base.*check"))
        if check_row is not None:
            self.check_zero_row(sensitivity_values_ws, check_row, "SENSITIVITY_CENTER_MISMATCH", "Sensitivity center check")
            return

        if not sheet_contains_any(sensitivity_ws, (r"\bbase\b", r"base case")):
            self.add_issue(
                "error",
                "SENSITIVITY_CENTER_MISSING",
                "Sensitivity sheet must mark the base-case center cell or provide a center check row.",
                "$.sheets.Sensitivity",
            )

    def check_company_type_rules(self) -> None:
        dcf_ws = self.sheet("DCF")
        workbook_text = " ".join(worksheet_text(self.wb[sheet]) for sheet in self.wb.sheetnames)

        if self.company_type in {"financial", "bank", "insurance", "broker"}:
            if dcf_ws is not None and sheet_contains_any(dcf_ws, (r"\bwacc\b", r"fcff", r"ufcf", r"free cash flow")):
                self.add_issue(
                    "error",
                    "FINANCIAL_FIRM_FCFF_WACC_DCF_USED",
                    "Financial firms must not use ordinary FCFF/WACC DCF; use equity-model methods instead.",
                    "$.sheets.DCF",
                    {"company_type": self.company_type},
                )

        if self.company_type in {"biopharma", "pre_revenue_biopharma", "innovative_drug"}:
            has_rnpv_or_sotp = any(
                normalize_token(sheet_name) in {"rnpv", "sotp"} or "rnpv" in normalize_token(sheet_name) or "sotp" in normalize_token(sheet_name)
                for sheet_name in self.wb.sheetnames
            )
            if not has_rnpv_or_sotp and matches_any(workbook_text, TARGET_PRICE_PATTERNS):
                self.add_issue(
                    "error",
                    "BIOPHARMA_TARGET_OUTPUT_WITHOUT_RNPV_SOTP",
                    "Pipeline-driven biopharma must use rNPV/SOTP before any target-price-style output.",
                    "$.workbook",
                    {"company_type": self.company_type},
                )

    def resolve_dcf_status(self, requested_status: str) -> str:
        requested_status = normalize_token(requested_status)
        if requested_status != "auto":
            return requested_status

        workbook_text = " ".join(worksheet_text(self.wb[sheet]) for sheet in self.wb.sheetnames).lower()
        if "dcf_applicability" in workbook_text and "disabled" in workbook_text:
            return "disabled"
        if self.has_sheet("DCF"):
            return "allowed"
        return "not_selected"

    def has_sheet(self, sheet_name: str) -> bool:
        return sheet_name in self.wb.sheetnames

    def sheet(self, sheet_name: str) -> Optional[Any]:
        return self.wb[sheet_name] if sheet_name in self.wb.sheetnames else None

    def values_sheet(self, sheet_name: str) -> Optional[Any]:
        return self.values_wb[sheet_name] if sheet_name in self.values_wb.sheetnames else None

    def find_source_columns(self, ws: Any) -> List[int]:
        source_columns = set()
        for row in range(1, min(ws.max_row, 15) + 1):
            for col in range(1, ws.max_column + 1):
                token = normalize_token(ws.cell(row=row, column=col).value)
                if token in {"source", "sourceid", "sourceids", "source_id"}:
                    source_columns.add(col)
        return sorted(source_columns)

    def add_issue(
        self,
        severity: str,
        code: str,
        message: str,
        path: str,
        details: Optional[Dict[str, Any]] = None,
    ) -> None:
        self.issues.append(Issue(severity, code, message, path, details or {}))

    def result(self) -> Dict[str, Any]:
        errors = [issue for issue in self.issues if issue.severity == "error"]
        warnings = [issue for issue in self.issues if issue.severity == "warning"]
        missing_sheets = [sheet for sheet in self.required_sheets if not self.has_sheet(sheet)]
        passed = not errors
        return {
            "validator": "model_validator",
            "validator_version": 1,
            "input_path": str(self.workbook_path),
            "passed": passed,
            "model_validation_status": "passed" if passed else "failed",
            "report_generation_allowed": passed,
            "dcf_status": self.dcf_status,
            "company_type": self.company_type,
            "model_blocked_reason": [
                {"code": issue.code, "message": issue.message, "path": issue.path}
                for issue in errors
            ],
            "summary": {
                "sheets_total": len(self.wb.sheetnames),
                "required_sheets": len(self.required_sheets),
                "missing_required_sheets": len(missing_sheets),
                "errors": len(errors),
                "warnings": len(warnings),
            },
            "required_sheets": self.required_sheets,
            "missing_required_sheets": missing_sheets,
            "issues": [asdict(issue) for issue in self.issues],
        }


def row_label(ws: Any, row: int) -> str:
    parts = []
    for col in range(1, min(ws.max_column, 4) + 1):
        value = ws.cell(row=row, column=col).value
        if value is not None:
            parts.append(str(value))
    return " ".join(parts).strip()


def row_has_numeric_data(ws: Any, row: int, excluded_columns: Iterable[int]) -> bool:
    excluded = set(excluded_columns)
    for col in range(1, ws.max_column + 1):
        if col in excluded:
            continue
        if numeric_value(ws.cell(row=row, column=col).value) is not None:
            return True
    return False


def row_has_source_id(ws: Any, row: int, source_columns: Iterable[int]) -> bool:
    for col in source_columns:
        if looks_like_source_id(ws.cell(row=row, column=col).value):
            return True
    return False


def looks_like_source_id(value: Any) -> bool:
    if is_blank(value) or is_formula(value):
        return False
    text = str(value).strip()
    token = normalize_token(text)
    if token in {"missing", "na", "n/a"}:
        return True
    return bool(re.search(r"(source|src|cninfo|edgar|hkex|sec|ifind|yahoo|filing|annual|10k|10q)", token))


def period_columns(ws: Any, historical: bool) -> List[int]:
    columns = []
    for row in range(1, min(ws.max_row, 8) + 1):
        for col in range(1, ws.max_column + 1):
            text = str(ws.cell(row=row, column=col).value or "").strip().lower()
            if not text:
                continue
            if historical:
                if re.search(r"(fy[- ]?\d|fy0|latest|actual|\ba\b|20\d{2}a)", text):
                    columns.append(col)
            else:
                if re.search(r"(fy\+\d|forecast|estimate|\be\b|20\d{2}e)", text):
                    columns.append(col)
    return sorted(set(columns))


def value_columns(ws: Any) -> List[int]:
    columns = []
    for col in range(3, ws.max_column + 1):
        for row in range(1, min(ws.max_row, 20) + 1):
            if not is_blank(ws.cell(row=row, column=col).value):
                columns.append(col)
                break
    return columns


def find_row(ws: Any, patterns: Sequence[str]) -> Optional[int]:
    for row in range(1, ws.max_row + 1):
        if matches_any(row_label(ws, row), patterns):
            return row
    return None


def find_label_numeric_value(
    ws: Optional[Any],
    patterns: Sequence[str],
    reject_patterns: Sequence[str] = (),
) -> Optional[float]:
    if ws is None:
        return None
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            text = str(ws.cell(row=row, column=col).value or "")
            if reject_patterns and matches_any(text, reject_patterns):
                continue
            if not matches_any(text, patterns):
                continue
            for candidate_col in range(col + 1, min(ws.max_column, col + 8) + 1):
                value = numeric_value(ws.cell(row=row, column=candidate_col).value)
                if value is not None:
                    return value
            for candidate_row in range(row + 1, min(ws.max_row, row + 3) + 1):
                value = numeric_value(ws.cell(row=candidate_row, column=col).value)
                if value is not None:
                    return value
    return None


def worksheet_text(ws: Any) -> str:
    parts = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                parts.append(str(cell.value))
    return " ".join(parts)


def sheet_contains_any(ws: Any, patterns: Sequence[str]) -> bool:
    return matches_any(worksheet_text(ws), patterns)


def formula_references_sheet(value: Any, sheet_name: str) -> bool:
    if not is_formula(value):
        return False
    normalized_formula = str(value).lower().replace("'", "")
    normalized_sheet = sheet_name.lower()
    return f"{normalized_sheet}!" in normalized_formula


def is_formula(value: Any) -> bool:
    return isinstance(value, str) and value.startswith("=")


def is_blank(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        return value.strip() == ""
    return False


def numeric_value(value: Any) -> Optional[float]:
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        text = value.strip()
        if not text or text.startswith("="):
            return None
        text = text.replace(",", "").replace("%", "")
        try:
            number = float(text)
        except ValueError:
            return None
        return number / 100.0 if "%" in value else number
    return None


def is_check_or_tie_label(label: str) -> bool:
    return matches_any(label, (r"check", r"tie-out", r"tie out"))


def matches_any(text: str, patterns: Sequence[str]) -> bool:
    normalized = str(text or "").lower()
    return any(re.search(pattern, normalized, re.IGNORECASE) for pattern in patterns)


def normalize_token(value: Any) -> str:
    return re.sub(r"[^a-zA-Z0-9_]+", "", str(value or "").strip().lower())


def cell_ref(ws: Any, row: int, col: int) -> str:
    return f"{ws.title}!{column_letter(col)}{row}"


def column_letter(col: int) -> str:
    if get_column_letter is not None:
        return str(get_column_letter(col))
    result = ""
    while col:
        col, remainder = divmod(col - 1, 26)
        result = chr(65 + remainder) + result
    return result


def build_failure_result(path: Path, issue: Issue) -> Dict[str, Any]:
    return {
        "validator": "model_validator",
        "validator_version": 1,
        "input_path": str(path),
        "passed": False,
        "model_validation_status": "failed",
        "report_generation_allowed": False,
        "dcf_status": "unknown",
        "company_type": "unknown",
        "model_blocked_reason": [
            {"code": issue.code, "message": issue.message, "path": issue.path}
        ],
        "summary": {
            "sheets_total": 0,
            "required_sheets": len(BASE_REQUIRED_SHEETS),
            "missing_required_sheets": len(BASE_REQUIRED_SHEETS),
            "errors": 1 if issue.severity == "error" else 0,
            "warnings": 1 if issue.severity == "warning" else 0,
        },
        "required_sheets": list(BASE_REQUIRED_SHEETS),
        "missing_required_sheets": list(BASE_REQUIRED_SHEETS),
        "issues": [asdict(issue)],
    }


def parse_args(argv: Optional[Sequence[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Validate an equity research Excel financial model.")
    parser.add_argument("--workbook", required=True, help="Path to .xlsx financial model workbook.")
    parser.add_argument(
        "--dcf-status",
        choices=sorted(VALID_DCF_STATUSES),
        default="auto",
        help="DCF applicability status from the valuation method router.",
    )
    parser.add_argument(
        "--company-type",
        choices=sorted(VALID_COMPANY_TYPES),
        default="general",
        help="Company/model type used for industry-specific valuation restrictions.",
    )
    parser.add_argument("--pretty", action="store_true", help="Pretty-print validation result JSON.")
    return parser.parse_args(argv)


def main(argv: Optional[Sequence[str]] = None) -> int:
    args = parse_args(argv)
    workbook_path = Path(args.workbook).resolve()

    if load_workbook is None:
        result = build_failure_result(
            workbook_path,
            Issue(
                "error",
                "OPENPYXL_UNAVAILABLE",
                "openpyxl is required to validate .xlsx workbooks. Install openpyxl or run in an environment that provides it.",
                "$.environment.openpyxl",
            ),
        )
    elif workbook_path.suffix.lower() != ".xlsx":
        result = build_failure_result(
            workbook_path,
            Issue("error", "UNSUPPORTED_WORKBOOK_FORMAT", "model_validator.py supports .xlsx files only.", str(workbook_path)),
        )
    elif not workbook_path.exists():
        result = build_failure_result(
            workbook_path,
            Issue("error", "WORKBOOK_NOT_FOUND", f"Workbook does not exist: {workbook_path}.", str(workbook_path)),
        )
    else:
        try:
            formula_wb = load_workbook(workbook_path, data_only=False)
            values_wb = load_workbook(workbook_path, data_only=True)
        except Exception as exc:
            result = build_failure_result(
                workbook_path,
                Issue("error", "WORKBOOK_LOAD_FAILED", str(exc), str(workbook_path)),
            )
        else:
            result = ModelValidator(
                formula_wb,
                values_wb,
                workbook_path,
                dcf_status=args.dcf_status,
                company_type=args.company_type,
            ).validate()

    print(json.dumps(result, ensure_ascii=False, indent=2 if args.pretty else None))
    return 0 if result["passed"] else 1


if __name__ == "__main__":
    sys.exit(main())
